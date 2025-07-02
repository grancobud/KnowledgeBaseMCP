"""
XLSX file creation and manipulation utilities
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import asyncio
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

logger = logging.getLogger(__name__)

class XlsxWriter:
    """Class for creating and manipulating XLSX files"""
    
    def __init__(self):
        if not XLSX_AVAILABLE:
            logger.warning("XLSX libraries not available. Install with: pip install openpyxl pandas")
    
    async def create_workbook(self, data: Dict[str, Any], file_path: str, **kwargs) -> str:
        """
        Create a new XLSX workbook with multiple sheets
        
        Args:
            data: Dictionary where keys are sheet names and values are sheet data
            file_path: Output file path
            **kwargs: Additional options like styling, formatting
        """
        if not XLSX_AVAILABLE:
            return "Error: openpyxl and pandas libraries not available. Install with: pip install openpyxl pandas"
        
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create new workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Process each sheet
            for sheet_name, sheet_data in data.items():
                ws = wb.create_sheet(title=sheet_name)
                
                if isinstance(sheet_data, dict):
                    await self._write_dict_to_sheet(ws, sheet_data, **kwargs)
                elif isinstance(sheet_data, list):
                    await self._write_list_to_sheet(ws, sheet_data, **kwargs)
                elif hasattr(sheet_data, 'to_excel'):  # pandas DataFrame
                    await self._write_dataframe_to_sheet(ws, sheet_data, **kwargs)
                else:
                    # Convert to string and write as simple text
                    ws['A1'] = str(sheet_data)
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            
            return f"XLSX workbook created successfully: {file_path}"
            
        except Exception as e:
            logger.error(f"Error creating XLSX workbook: {str(e)}")
            return f"Error creating XLSX workbook: {str(e)}"
    
    async def create_dataframe_workbook(self, dataframes: Dict[str, Any], file_path: str, **kwargs) -> str:
        """
        Create XLSX workbook from pandas DataFrames
        
        Args:
            dataframes: Dictionary of sheet_name -> DataFrame
            file_path: Output file path
            **kwargs: Formatting options
        """
        if not XLSX_AVAILABLE:
            return "Error: openpyxl and pandas libraries not available. Install with: pip install openpyxl pandas"
        
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=kwargs.get('include_index', True))
                    
                    # Apply formatting if requested
                    if kwargs.get('apply_formatting', True):
                        worksheet = writer.sheets[sheet_name]
                        await self._apply_default_formatting(worksheet)
            
            return f"DataFrame workbook created successfully: {file_path}"
            
        except Exception as e:
            logger.error(f"Error creating DataFrame workbook: {str(e)}")
            return f"Error creating DataFrame workbook: {str(e)}"
    
    async def append_to_workbook(self, file_path: str, sheet_name: str, data: Any, **kwargs) -> str:
        """
        Append data to existing workbook or create new one
        
        Args:
            file_path: XLSX file path
            sheet_name: Target sheet name
            data: Data to append (list, dict, DataFrame)
            **kwargs: Additional options
        """
        if not XLSX_AVAILABLE:
            return "Error: openpyxl and pandas libraries not available. Install with: pip install openpyxl pandas"
        
        try:
            path = Path(file_path)
            
            # Load existing workbook or create new one
            if path.exists():
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
                # Remove default sheet
                wb.remove(wb.active)
            
            # Get or create sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find next available row
                next_row = ws.max_row + 1
            else:
                ws = wb.create_sheet(title=sheet_name)
                next_row = 1
            
            # Write data starting from next_row
            if isinstance(data, list):
                await self._append_list_to_sheet(ws, data, next_row, **kwargs)
            elif isinstance(data, dict):
                await self._append_dict_to_sheet(ws, data, next_row, **kwargs)
            elif hasattr(data, 'to_excel'):  # pandas DataFrame
                await self._append_dataframe_to_sheet(ws, data, next_row, **kwargs)
            else:
                ws.cell(row=next_row, column=1, value=str(data))
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            
            return f"Data appended to {file_path}, sheet '{sheet_name}'"
            
        except Exception as e:
            logger.error(f"Error appending to workbook: {str(e)}")
            return f"Error appending to workbook: {str(e)}"
    
    async def create_report_workbook(self, report_data: Dict[str, Any], file_path: str) -> str:
        """
        Create a formatted report workbook with multiple sections
        
        Args:
            report_data: Structured report data
            file_path: Output file path
        """
        if not XLSX_AVAILABLE:
            return "Error: openpyxl and pandas libraries not available. Install with: pip install openpyxl pandas"
        
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            
            # Create summary sheet
            summary_ws = wb.active
            summary_ws.title = "Summary"
            
            # Add report title and metadata
            summary_ws['A1'] = report_data.get('title', 'Report')
            summary_ws['A1'].font = Font(size=16, bold=True)
            summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            summary_ws['A3'] = report_data.get('description', '')
            
            # Create data sheets
            data_sections = report_data.get('data', {})
            for section_name, section_data in data_sections.items():
                ws = wb.create_sheet(title=section_name)
                
                if isinstance(section_data, dict) and 'dataframe' in section_data:
                    # Handle DataFrame with metadata
                    df = section_data['dataframe']
                    await self._write_dataframe_to_sheet(ws, df, apply_formatting=True)
                elif isinstance(section_data, list):
                    await self._write_list_to_sheet(ws, section_data, apply_formatting=True)
                else:
                    await self._write_dict_to_sheet(ws, section_data, apply_formatting=True)
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            
            return f"Report workbook created successfully: {file_path}"
            
        except Exception as e:
            logger.error(f"Error creating report workbook: {str(e)}")
            return f"Error creating report workbook: {str(e)}"
    
    async def _write_dict_to_sheet(self, worksheet, data: Dict, **kwargs):
        """Write dictionary data to worksheet"""
        row = 1
        for key, value in data.items():
            worksheet.cell(row=row, column=1, value=str(key))
            worksheet.cell(row=row, column=2, value=str(value))
            row += 1
        
        if kwargs.get('apply_formatting', False):
            await self._apply_default_formatting(worksheet)
    
    async def _write_list_to_sheet(self, worksheet, data: List, **kwargs):
        """Write list data to worksheet"""
        for row_idx, item in enumerate(data, 1):
            if isinstance(item, (list, tuple)):
                for col_idx, value in enumerate(item, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=str(value))
            elif isinstance(item, dict):
                # First row as headers if first item is dict
                if row_idx == 1:
                    for col_idx, key in enumerate(item.keys(), 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=str(key))
                    row_idx += 1
                for col_idx, value in enumerate(item.values(), 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=str(value))
            else:
                worksheet.cell(row=row_idx, column=1, value=str(item))
        
        if kwargs.get('apply_formatting', False):
            await self._apply_default_formatting(worksheet)
    
    async def _write_dataframe_to_sheet(self, worksheet, df, **kwargs):
        """Write pandas DataFrame to worksheet"""
        # Write DataFrame to worksheet
        for r in dataframe_to_rows(df, index=kwargs.get('include_index', True), header=True):
            worksheet.append(r)
        
        if kwargs.get('apply_formatting', True):
            await self._apply_default_formatting(worksheet)
    
    async def _append_list_to_sheet(self, worksheet, data: List, start_row: int, **kwargs):
        """Append list data to worksheet starting from start_row"""
        current_row = start_row
        for item in data:
            if isinstance(item, (list, tuple)):
                for col_idx, value in enumerate(item, 1):
                    worksheet.cell(row=current_row, column=col_idx, value=str(value))
            elif isinstance(item, dict):
                for col_idx, value in enumerate(item.values(), 1):
                    worksheet.cell(row=current_row, column=col_idx, value=str(value))
            else:
                worksheet.cell(row=current_row, column=1, value=str(item))
            current_row += 1
    
    async def _append_dict_to_sheet(self, worksheet, data: Dict, start_row: int, **kwargs):
        """Append dictionary data to worksheet starting from start_row"""
        current_row = start_row
        for key, value in data.items():
            worksheet.cell(row=current_row, column=1, value=str(key))
            worksheet.cell(row=current_row, column=2, value=str(value))
            current_row += 1
    
    async def _append_dataframe_to_sheet(self, worksheet, df, start_row: int, **kwargs):
        """Append DataFrame data to worksheet starting from start_row"""
        # Convert DataFrame to rows and append
        for r_idx, r in enumerate(dataframe_to_rows(df, index=kwargs.get('include_index', True), header=False)):
            for c_idx, value in enumerate(r, 1):
                worksheet.cell(row=start_row + r_idx, column=c_idx, value=value)
    
    async def _apply_default_formatting(self, worksheet):
        """Apply default formatting to worksheet"""
        try:
            # Header formatting (first row)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Apply header formatting to first row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
                        
        except Exception as e:
            logger.warning(f"Error applying formatting: {str(e)}")
